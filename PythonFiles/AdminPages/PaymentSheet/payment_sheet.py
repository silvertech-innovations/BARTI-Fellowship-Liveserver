from datetime import date, timedelta, datetime
from io import BytesIO
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
from classes.database import HostConfig, ConfigPaths, ConnectParam
from fpdf import FPDF
from flask import Blueprint, render_template, session, request, redirect, url_for, flash, make_response

payment_sheet_blueprint = Blueprint('payment_sheet', __name__)


def payment_sheet_auth(app):
    # ------ HOST Configs are in classes/connection.py
    host = HostConfig.host
    app_paths = ConfigPaths.paths.get(host)
    if app_paths:
        for key, value in app_paths.items():
            app.config[key] = value

    @payment_sheet_blueprint.route('/payment_sheet', methods=['GET', 'POST'])
    def payment_sheet():
        if not session.get('logged_in'):
            # Redirect to the admin login page if the user is not logged in
            return redirect(url_for('admin_login'))
        user_records = []
        if request.method == 'GET':
            # Establish a database connection
            host = HostConfig.host
            connect_param = ConnectParam(host)
            cnx, cursor = connect_param.connect(use_dict=True)
            # print('I have made connection')

            # Fetch user data based on the email
            cursor.execute("""

                        SELECT * 
                        FROM application_page 
                        WHERE final_approval = 'accepted' 
                          AND phd_registration_year >= '2023'

                        UNION

                        SELECT * 
                        FROM application_page 
                        WHERE phd_registration_year > '2020' 
                          AND aadesh = 1;

            """)
            user_data = cursor.fetchall()  # Use fetchall to retrieve all rows
            # print('user data:', user_data)

            for row in user_data:
                # Calculate values based on user data
                applicant_id = row['applicant_id']
                faculty = row["faculty"]
                # print('faculty', faculty)
                fellowship_awarded_year = row['fellowship_awarded_year']
                fellowship_awarded_date = row['fellowship_awarded_date']
                joining_date = row["phd_registration_date"]
                city = row['research_center_district']
                bank_name = row['bank_name']
                account_number = row['account_number']
                ifsc = row['ifsc_code']
                year = '2023'
                # print(joining_date)

                # Calculate Count Yearly
                if faculty == "Arts":
                    count_yearly = 20500
                elif faculty == "Law":
                    count_yearly = 20500
                elif faculty == "Commerce":
                    count_yearly = 20500
                elif faculty == "Other":
                    count_yearly = 20500
                elif faculty == "Science":
                    count_yearly = 25000
                else:
                    count_yearly = 0  # Handle other faculty values as needed


                if city in [
                            'Hyderabad(UA)', 'Delhi(UA)', 'Ahmadabad(UA)', 'Bengalore / Bengaluru(UA)',
                            'Greater Mumbai(UA)', 'Pune(UA)', 'Chennai(UA)', 'Kolkata(UA)'
                            ]:
                    rate = '30%'
                elif city in [
                            "Vijayawada (UA)", "Warangal (UA)", "Greater Visakhapatnam (M.Corpn.)", "Guntur (UA)",
                            "Nellore (UA)", "Guwahati (UA)", "Patna (UA)", "Chandigarh (UA)",
                            "Durg-Bhilai Nagar (UA)", "Raipur (UA)", "Rajkot (UA)", "Jamnagar (UA)",
                            "Bhavnagar (UA)", "Vadodara (UA)", "Surat (UA)", "Faridabad (M.Corpn.)",
                            "Gurgaon (UA)", "Srinagar (UA)", "Jammu (UA)", "Jamshedpur (UA)",
                            "Dhanbad (UA)", "Ranchi (UA)", "Bokaro Steel City (UA)", "Belgaum (UA)",
                            "Hubli-Dharwad (M.Corpn.)", "Mangalore (UA)", "Mysore (UA)", "Gulbarga (UA)",
                            "Kozhikode (UA)", "Kochi (UA)", "Thiruvananthapuram (UA)", "Thrissur (UA)",
                            "Malappuram (UA)", "Kannur (UA)", "Kollam (UA)", "Gwalior (UA)",
                            "Indore (UA)", "Bhopal (UA)", "Jabalpur (UA)", "Ujjain (M.Corpn.)",
                            "Amravati (M.Corpn.)", "Nagpur (UA)", "Aurangabad (UA)", "Nashik (UA)",
                            "Bhiwandi (UA)", "Solapur (M.Corpn.)", "Kolhapur (UA)", "Vasai-Virar City (M.Corpn.)",
                            "Malegaon (UA)", "Nanded-Waghala (M. Corpn.)", "Sangli (UA)", "Cuttack (UA)",
                            "Bhubaneswar (UA)", "Raurkela (UA)", "Puducherry/Pondicherry (UA)", "Amritsar (UA)",
                            "Jalandhar (UA)", "Ludhiana (M.Corpn.)", "Bikaner (M.Corpn.)", "Jaipur (M.Corpn.)",
                            "Jodhpur (UA)", "Kota (M.Corpn.)", "Ajmer (UA)", "Salem (UA)",
                            "Tiruppur (UA)", "Coimbatore (UA)", "Tiruchirappalli (UA)", "Madurai (UA)",
                            "Erode (UA)", "Moradabad (M.Corpn.)", "Meerut (UA)", "Ghaziabad (UA)",
                            "Aligarh (UA)", "Agra (UA)", "Bareilly (UA)", "Lucknow (UA)",
                            "Kanpur (UA)", "Allahabad (UA)", "Gorakhpur (UA)", "Varanasi (UA)",
                            "Saharanpur (M.Corpn.)", "Noida (CT)", "Firozabad (NPP)", "Jhansi (UA)",
                            "Dehradun (UA)", "Asansol (UA)", "Siliguri (UA)", "Durgapur (UA)"
                            ]:
                    rate = '20%'
                else:
                    rate = '10%'

                # print("Rate:", rate)

                # Initialize the "from" and "to" date to empty strings
                duration_date_from = ""
                duration_date_to = ""

                if joining_date:  # Check if joining_date is not None
                    # Calculate Duration Date (adding 3 months to joining date)
                    duration_date_from = fellowship_awarded_date  # Assuming this is a datetime object
                    duration_date_to = fellowship_awarded_date + timedelta(days=90)  # Adding 90 days to joining date
                    # Extract day, month, and year
                    day = duration_date_to.day
                    month = duration_date_to.month
                    year = duration_date_to.year

                    # Print the stripped day, month, and year
                    # print(f"Day: {day}, Month: {month}, Year: {year}")
                    # Format the dates for display in the desired format
                    duration_date_from_str = duration_date_from.strftime('%d/%m/%Y')  # "17 Aug 2023"
                    duration_date_to_str = duration_date_to.strftime('%d/%m/%Y')  # "15 Nov 2023"

                # Calculate Total Months
                total_months = 3

                # Calculate Fellowship
                fellowship = 42000  # Fixed value for 3 months

                # Calculate Total Fellowship
                total_fellowship = fellowship * total_months

                rate_str = float(rate.rstrip('%'))
                convert_rate = (rate_str / 100)
                hra_amount = convert_rate * fellowship

                months = total_months

                total_hra = hra_amount * months

                total = total_fellowship + total_hra

                # Calculate the date 2 years after duration_date_from
                two_years_later = duration_date_from + timedelta(days=730)  # 2 years = 730 days

                # Assuming 'phd_registration_date' is already a datetime object
                if 'phd_registration_date' in row and row['phd_registration_date']:
                    joiningDate = row['phd_registration_date'].strftime('%Y-%m-%d')

                # Get the current date
                current_date = datetime.now().date()
                current_year = current_date.year

                # Check the category based on 2 years difference
                if current_year == fellowship_awarded_year + 2:
                    category = "SRF"  # Senior Research Fellowship
                else:
                    category = "JRF"  # Junior Research Fellowship

                # Create a record dictionary for the user
                record = {
                    "applicant_id": row['applicant_id'],
                    "full_name": str(row['first_name']) + ' ' + str(row['middle_name']) + ' ' + str(row['last_name']),
                    "first_name": row['first_name'],
                    "last_name": row['last_name'],
                    "middle_name": row['middle_name'],
                    "email": row["email"],
                    "faculty": row['faculty'],
                    "fellowship_awarded_date": fellowship_awarded_date,
                    "joining_date": joiningDate,
                    "city": row['city'],
                    "duration": f"{duration_date_from_str} <span class='fw-bold'>to</span> {duration_date_to_str}",
                    "rate": rate,
                    "count": count_yearly,
                    "amount": hra_amount,
                    "months": months,
                    "total_hra": total_hra,
                    "total": total,
                    "duration_date_from": duration_date_from,
                    "duration_date_to": duration_date_to,
                    "duration_day": day,
                    "duration_month": month,
                    "duration_year": year,
                    "total_months": total_months,
                    "fellowship": fellowship,
                    "to_fellowship": total_fellowship,
                    "phd_registration_year": row['phd_registration_year'],
                    "id": row['id'],
                    "account_number": account_number,
                    "ifsc": ifsc,
                    "bank_name": bank_name,
                    "year": year,
                    "jrf_srf": category
                }

                user_records.append(record)

                email = record['email']

                host = HostConfig.host
                connect_param = ConnectParam(host)
                cnx, cursor = connect_param.connect()

                cursor.execute(" SELECT * FROM payment_sheet where email=%s", (email,))
                result = cursor.fetchone()

                if result:
                    pass
                    # print("Existing Record:", email)
                    # Record already exists, do not insert again
                else:
                    # print("Record not found, proceeding with the INSERT query")
                    # Insert values into the payment_sheet table
                    host = HostConfig.host
                    connect_param = ConnectParam(host)
                    cnx, cursor = connect_param.connect()

                    insert_query = """
                        INSERT INTO payment_sheet (
                            full_name, faculty, fellowship_awarded_date, city, date, jrf_srf, duration_date_from, duration_date_to, duration_day, duration_month, duration_year, 
                            rate, count, amount, months, total_hra, total,
                            total_months, fellowship,
                            to_fellowship, bank_name, ifsc_code, account_number, fellowship_awarded_year, email
                        )
                        VALUES (%(full_name)s, %(faculty)s, %(fellowship_awarded_date)s, %(city)s, %(joining_date)s, %(jrf_srf)s, %(duration_date_from)s, %(duration_date_to)s,
                                %(duration_day)s, %(duration_month)s, %(duration_year)s,
                                %(rate)s, %(count)s, %(amount)s, %(months)s, %(total_hra)s, %(total)s, 
                                %(total_months)s, %(fellowship)s, %(to_fellowship)s, %(bank_name)s, %(ifsc)s,
                                %(account_number)s, %(year)s, %(email)s)         
                    """
                    # Execute the INSERT query
                    cursor.execute(insert_query, record)

                    # Commit the changes to the database
                    cnx.commit()

                # Close the database cursor and connection
                cursor.close()
                cnx.close()
            # Close the database cursor and connection
            cursor.close()
            cnx.close()

        return render_template('AdminPages/PaymentSheet/payment_sheet.html', user_records=user_records)

    class PDF(FPDF):
        def __init__(self):
            super().__init__(orientation='P', unit='mm', format='A3')  # Set format to A3

        header_added = False  # To track whether the header is added to the first page

        def header(self):
            if not self.header_added:
                var = get_base_url()
                print(var)

                # Adjusted dimensions for A3 format
                self.set_font("Arial", "B", 12)

                # Adjust the X, Y, and image size to fit A3 format
                # for LOCALSERVER
                # self.image('/var/www/icswebapp/icswebapp/static/Images/satya.png', 140, 10,
                #            30)  # Adjusted position and size for A3
                # image_width = 140  # Updated to a size appropriate for A3
                # text_x_position = self.get_x()
                # text_y_position = self.get_y() + 25  # Adjusted for A3 format
                # self.set_xy(text_x_position, text_y_position)
                # # Adjusted positions for A3 format
                # self.image('/var/www/icswebapp/icswebapp/static/Images/newtrtiImage.png', 20, 10,
                #            60)  # Adjust size for larger format
                # self.image('/var/www/icswebapp/icswebapp/static/Images/mahashasn_new.png', 215, 10,
                #            60)  # Adjust size and position for A3

                # For HOSTSERVER
                self.image('static/Images/satya.png', 140, 10, 30)  # Adjusted position and size for A3
                image_width = 140  # Updated to a size appropriate for A3
                text_x_position = self.get_x()
                text_y_position = self.get_y() + 25  # Adjusted for A3 format
                self.set_xy(text_x_position, text_y_position)
                # Adjusted positions for A3 format
                self.image('static/Images/newtrtiImage.png', 20, 10, 60)  # Adjust size for larger format
                self.image('static/Images/mahashasn_new.png', 215, 10, 60)  # Adjust size and position for A3

                # Centered text for A3 format
                self.ln(10)
                self.cell(0, 10, "Government of Maharashtra", align="C", ln=True)
                self.cell(0, 10, "Tribal Research & Training Institute", align="C", ln=True)
                self.cell(0, 10, "28, Queens Garden, Pune - 411001", align="C", ln=True)

                # Adjust the dashed line width for A3
                self.dashed_line(10, self.get_y(), 290, self.get_y(), dash_length=3,
                                 space_length=1)  # Adjust for A3 width

                self.ln(5)  # Adjust space after the line
                self.set_font("Arial", size=10)

                current_date = datetime.now().strftime('%B %Y')
                # Left-aligned and right-aligned text for A3 format
                self.cell(0, 10, "No.: Research-_____/Case.No ____/Desk- __/_____ ", ln=False)
                self.cell(0, 10, f"Date: {current_date}", align="R", ln=True)

                self.set_font("Arial", "B", size=12)
                self.cell(0, 10, "Appendix", align="C", ln=True)

                # Adjust rotation for A3
                # self.ln(2)
                # self.rotate(45)
                # self.set_font('Arial', '', 65)  # Adjust font size for A3
                # self.set_text_color(192, 192, 192)

                # Adjust the position for the watermark on A3
                # self.text(-50, 280, "STRF FELLOWSHIP")  # Adjust position for A3 layout
                # self.rotate(0)

                self.header_added = True  # Set to True after adding the header

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", 'I', 8)
            self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

    def get_base_url():
        base_url = request.url_root
        return base_url

    @payment_sheet_blueprint.route('/export_payment_sheet_pdf')
    def export_payment_sheet_pdf():
        # Establish a database connection
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        cursor.execute("SELECT number, full_name, email, faculty, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
                       "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
                       "city, bank_name, ifsc_code, account_number FROM payment_sheet")

        data = cursor.fetchall()
        pdf = PDF()
        pdf.add_page()

        # Set margins
        pdf.set_margins(5, 5, 10)  # Left, Top, Right margins

        current_date = datetime.now().strftime('%B %Y')

        # Add date information
        pdf.set_font("Arial", 'I', 12)
        # pdf.cell(0, 10, f'Number:', ln=True, align='L')
        # pdf.cell(0, 10, f'Date: {current_date}', ln=True, align='L')
        pdf.ln(5)  # Small line break

        # Set header
        pdf.set_font("Arial", 'B', 10)

        # Define fixed column widths
        column_widths = [15, 60, 30, 45, 40, 40, 30, 30]  # Set fixed widths for each column
        headers = ['Sr. No.', 'Name of Student', 'Date of PHD Reg.', 'Fellowship Awarded Date', 'Duration', 'Bank Name',
                   'Account Number', 'IFSC', 'Amount']

        # Add header row with multi-cell for text wrapping
        x_start = pdf.get_x()  # Get the starting x position

        for i, header in enumerate(headers):
            # Store the current x and y position before adding the multi_cell
            x = pdf.get_x()
            y = pdf.get_y()

            # Add the multi_cell for the header
            pdf.cell(column_widths[i], 10, header, border=1, align='C')

            # Set the position back to where it was before adding multi_cell for the next cell
            pdf.set_xy(x + column_widths[i], y)

        pdf.ln()  # Move to the next line after adding the header

        # Set font for data
        pdf.set_font("Arial", '', 10)

        # Add data to the PDF
        for index, row in enumerate(data, start=1):
            pdf.cell(column_widths[0], 10, str(index), 1, align='C')

            # Add the name cell with multi_cell
            full_name = row['full_name']
            pdf.cell(column_widths[1], 10, full_name, 1, align='C')

            # Move cursor back to the right for the next column
            # pdf.cell(column_widths[0], 10, '', 0)  # Placeholder for Sr. No.

            # Add the remaining cells
            pdf.cell(column_widths[2], 10, str(row['date']), 1, align='C')
            duration_text = f"{row['duration_date_from']} to {row['duration_date_to']}"
            pdf.cell(column_widths[3], 10, duration_text, 1, align='C')
            pdf.cell(column_widths[4], 10, row['bank_name'] if row['bank_name'] else 'N/A', 1, align='C')
            pdf.cell(column_widths[5], 10, row['account_number'] if row['account_number'] else 'N/A', 1, align='C')
            pdf.cell(column_widths[6], 10, row['ifsc_code'] if row['ifsc_code'] else 'N/A', 1, align='C')
            pdf.cell(column_widths[7], 10, str(row['total']), 1, align='C')

            # Move to the next line
            pdf.ln()

        # Finalize PDF output
        response = make_response(pdf.output(dest='S').encode('latin1'))
        response.headers['Content-Disposition'] = 'attachment; filename=Payment_Sheet_2023_2024.pdf'
        response.headers['Content-Type'] = 'application/pdf'

        return response

    @payment_sheet_blueprint.route('/export_payment_sheet')
    def export_payment_sheet():
        # Establish a database connection
        host = HostConfig.host
        connect_param = ConnectParam(host)
        cnx, cursor = connect_param.connect(use_dict=True)

        cursor.execute("SELECT number, full_name, email, faculty, fellowship_awarded_date, date, duration_date_from, duration_date_to, "
                       "total_months, fellowship, to_fellowship, rate, amount, months, total_hra, count, pwd, total,"
                       "city, bank_name, ifsc_code, account_number FROM payment_sheet")

        data = cursor.fetchall()
        print(data)

        # Create a workbook and add a worksheet
        wb = Workbook()
        ws = wb.active

        header = 4
        date_column_index = 1
        date_column_index_date = 8
        current_date = datetime.now().strftime('%B %Y')
        # Add the bold row before the header
        ws.cell(row=1, column=header, value='Appendix')  # Place "Date:" in column B
        ws.cell(row=2, column=date_column_index, value='Number:')  # Place "Date:" in column B
        ws.cell(row=2, column=date_column_index_date, value=f'Date:{current_date}')  # Place "Date:" in column B
        # ws.cell(row=1, column=date_column_index_date + 1, value=current_date)  # Place the date in column C
        # Add the bold row before the header
        # ws.append(['Date:', current_date])  # Format as "Date: September 2024"
        bold_row = ws[1]  # Get the last added row (the one we just added)
        bold_row_2 = ws[2]  # Get the last added row (the one we just added)
        for cell in bold_row and bold_row_2:
            cell.font = Font(bold=True)  # Make the text bold

        # Add header row
        ws.append(['Sr. No.', 'Name of Student', 'Date of PHD Registration', 'Fellowship Awarded Date',
                   'Duration', 'Bank Name', 'Account Number', 'IFSC', 'Fellowship Amount'])

        # Add data to the worksheet with formatting
        for index, row in enumerate(data, start=1):
            full_name = row['full_name']

            # Joining date
            joining_date = row['date']
            fellowship_awarded_date = row['fellowship_awarded_date']
            # Duration dates
            duration_date_from = row['duration_date_from']
            duration_date_to = row['duration_date_to']

            # Convert string dates to datetime objects if they are not already
            if isinstance(duration_date_from, str):
                try:
                    duration_date_from = datetime.strptime(duration_date_from, '%Y-%m-%d')  # Adjust format as needed
                except ValueError:
                    duration_date_from = None

            if isinstance(duration_date_to, str):
                try:
                    duration_date_to = datetime.strptime(duration_date_to, '%Y-%m-%d')  # Adjust format as needed
                except ValueError:
                    duration_date_to = None

            if isinstance(duration_date_from, datetime) and isinstance(duration_date_to, datetime):
                duration_date_from_str = duration_date_from.strftime('%d %b %Y')  # Format as "17 Aug 2023"
                duration_date_to_str = duration_date_to.strftime('%d %b %Y')  # Format as "15 Nov 2023"
                duration = f"{duration_date_from_str} to {duration_date_to_str}"
            else:
                duration = "N/A"

            # Other fields
            bank_name = row['bank_name']
            account_number = row['account_number']
            ifsc = row['ifsc_code']
            fellowship_amount = row['total']

            # Append the formatted data
            ws.append([index, full_name, joining_date, fellowship_awarded_date, duration,
                       bank_name, account_number, ifsc, fellowship_amount])

        # Save the workbook in memory as bytes
        data = BytesIO()
        wb.save(data)
        data.seek(0)

        # Create a response object and attach the workbook as a file
        response = make_response(data.getvalue())
        response.headers['Content-Disposition'] = 'attachment; filename=Payment_Sheet_2023_2024.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        return response